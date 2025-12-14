import sys
import csv
from datetime import datetime
from pathlib import Path
from PyQt6.QtWidgets import (QApplication, QMainWindow, QTableView, QVBoxLayout,
                             QWidget, QPushButton, QMessageBox, QHBoxLayout, QDialog,
                             QTabWidget, QLabel, QDateEdit, QComboBox, QGridLayout,
                             QFrame, QTextEdit, QMenuBar, QFileDialog, QGroupBox, QButtonGroup,
                             QLineEdit, QInputDialog, QRadioButton, QDialogButtonBox)
from PyQt6.QtSql import QSqlDatabase, QSqlQueryModel, QSqlQuery
from PyQt6.QtCore import Qt, QDate, QTimer
from PyQt6.QtGui import QAction, QIcon
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

from views.asset_dialog import AssetDialog
from views.issue_dialog import IssueDialog
from views.return_dialog import ReturnDialog
from views.edit_asset_dialog import EditAssetDialog
from views.login_dialog import LoginDialog
from views.request_dialog import RequestAssetDialog
from database.db_manager import DatabaseManager
from notification_manager import NotificationManager
from theme_manager import ThemeManager

from audit_logger import AuditLogger


class MainWindow(QMainWindow):
    def __init__(self, current_user=None):
        super().__init__()
        print("Инициализация главного окна...")
        self.db = DatabaseManager()
        
        # Сохраняем информацию о текущем пользователе
        self.current_user = current_user or {
            'user_id': 0,
            'username': 'guest',
            'role': 'guest',
            'employee_id': None,
            'full_name': 'Guest'
        }
        
        # Устанавливаем иконку окна (дублируем для надежности)
        try:
            logo_path = Path("pictures/InstrumentTracker_logo.png")
            if logo_path.exists():
                self.setWindowIcon(QIcon(str(logo_path)))
        except Exception as e:
            print(f"️ Не удалось установить иконку окна: {e}")
        
        # Инициализация менеджера тем
        self.current_theme = ThemeManager.load_theme()
        self.apply_theme(self.current_theme)
        
        # Инициализация менеджера уведомлений
        self.notification_manager = NotificationManager(self)
        self.notification_manager.start_checking(interval_ms=60000)  # Проверка каждую минуту
        
        # ВРЕМЕННО: Настройка email для тестирования
        # TODO: Добавить настройку через интерфейс
        # self.notification_manager.configure_email('your-email@yandex.ru', 'your-app-password')
        # self.notification_manager.start_email_checking(interval_ms=3600000)  # Проверка раз в час
        self.notification_manager.configure_email('AlexAndreev132@yandex.ru', 'пароль_приложения')
        self.notification_manager.start_email_checking(interval_ms=3600000)  # раз в час

        # Флаг для показа уведомлений при первом показе окна
        self._startup_notifications_shown = False

        self.init_ui()

    def _run_startup_notifications(self):
        """Показать приветственное уведомление и проверить уведомления пользователя."""
        try:
            # Приветственное сообщение при каждом запуске (автозакрывается)
            full_name = self.current_user.get('full_name') if self.current_user else None
            if not full_name and self.current_user and self.current_user.get('employee_id'):
                # Попробуем получить ФИО из БД
                try:
                    emp = self.db.execute_query("SELECT last_name || ' ' || first_name || COALESCE(' ' || patronymic, '') as full_name FROM Employees WHERE employee_id = ?", (self.current_user.get('employee_id'),))
                    if emp and emp[0][0]:
                        full_name = emp[0][0]
                except Exception:
                    full_name = None

            greeting = f"Добро пожаловать, {full_name}!" if full_name else "Добро пожаловать!"
            print(f"Показ приветственного уведомления: {greeting}")
            self.notification_manager.show_notification('info', '', greeting, persistent=False)

            # Для обычных пользователей показываем просрочки и напоминания на завтра
            # Задержка 4.5 сек: приветствие автозакрывается через 4 сек, затем показываем уведомления
            if self.current_user and self.current_user.get('role') != 'admin':
                employee_id = self.current_user.get('employee_id')
                if employee_id:
                    print(f"Проверка уведомлений для сотрудника {employee_id}...")
                    QTimer.singleShot(4500, lambda: self.notification_manager.check_user_notifications(employee_id))
            # Для админов показываем уведомление о всех просрочках
            elif self.current_user and self.current_user.get('role') == 'admin':
                print("Проверка просрочек и запросов для админа...")
                QTimer.singleShot(4500, lambda: self.notification_manager.check_admin_overdue())
                QTimer.singleShot(5500, lambda: self.notification_manager.check_new_requests_for_admin())
        except Exception as e:
            print(f" Ошибка при показе стартовых уведомлений: {e}")
            import traceback
            traceback.print_exc()

    def showEvent(self, event):
        """Вызывается при показе окна - показываем уведомления при первом входе"""
        super().showEvent(event)
        if not self._startup_notifications_shown:
            self._startup_notifications_shown = True
            print("Окно отображается, показываем стартовые уведомления...")
            QTimer.singleShot(300, self._run_startup_notifications)

    def apply_theme(self, theme_name):
        """Применить тему к приложению"""
        try:
            theme = ThemeManager.get_theme(theme_name)
            self.setStyleSheet(theme['app_stylesheet'])
            self.current_theme = theme_name
            # Обновляем стиль user_info_label
            if hasattr(self, 'user_info_label'):
                self.user_info_label.setStyleSheet(theme['user_info_style'])
            print(f"Применена тема: {theme_name}")
        except Exception as e:
            print(f"Ошибка при применении темы: {e}")
    
    def set_theme(self, theme_name):
        """Установить и сохранить тему"""
        self.apply_theme(theme_name)
        ThemeManager.save_theme(theme_name)
        # Обновляем все активные уведомления
        if hasattr(self, 'notification_manager'):
            self.notification_manager.update_all_notifications_theme(theme_name)
        
        # Обновляем user_info_label с новым стилем
        theme = ThemeManager.get_theme(theme_name)
        if hasattr(self, 'user_info_label'):
            self.user_info_label.setStyleSheet(theme['user_info_style'])
        
        # Обновляем тему всех открытых уведомлений
        if hasattr(self, 'notification_manager'):
            self.notification_manager.update_notifications_theme()


    def init_ui(self):
        """Инициализация пользовательского интерфейса"""
        self.setWindowTitle("Система учета инструментов и расходников - АО КОНСИСТ-ОС")
        self.setGeometry(100, 100, 1200, 700)

        # Создаем меню
        self.create_menu()

        # Центральный виджет с вкладками
        central_widget = QWidget()
        self.setCentralWidget(central_widget)

        # Основной layout
        layout = QVBoxLayout(central_widget)
        
        # Информация о пользователе
        self.user_info_label = QLabel(f"👤 Вы вошли как: {self.current_user.get('full_name', 'Unknown')} ({self.current_user.get('role', 'user').upper()})")
        theme = ThemeManager.get_theme(self.current_theme)
        self.user_info_label.setStyleSheet(theme['user_info_style'])
        layout.addWidget(self.user_info_label)

        # Создаем вкладки
        self.tabs = QTabWidget()

        # Вкладка 0: Панель управления (дашборд)
        self.dashboard_tab = QWidget()
        self.setup_dashboard_tab()
        self.tabs.addTab(self.dashboard_tab, "🏠 Панель управления")

        # Вкладка 1: Каталог активов (доступ: админ)
        if self.current_user.get('role') == 'admin':
            self.assets_tab = QWidget()
            self.setup_assets_tab()
            self.tabs.addTab(self.assets_tab, "📋 Каталог активов")

        # Вкладка 2: Операции
        self.operations_tab = QWidget()
        self.setup_operations_tab()
        self.tabs.addTab(self.operations_tab, "🔄 Операции")

        # Вкладка 3: Запросы на выдачу (только для админа)
        if self.current_user.get('role') == 'admin':
            self.requests_tab = QWidget()
            self.setup_requests_tab()
            self.tabs.addTab(self.requests_tab, "📬 Запросы")
            
            # Вкладка 4: Аккаунты (только для админа)
            self.accounts_tab = QWidget()
            self.setup_accounts_tab()
            self.tabs.addTab(self.accounts_tab, "👥 Аккаунты")
        else:
            # Вкладка для обычного пользователя: Мой профиль
            self.user_profile_tab = QWidget()
            self.setup_user_profile_tab()
            self.tabs.addTab(self.user_profile_tab, "👤 Мой профиль")

        # Вкладка 5: Отчеты
        self.reports_tab = QWidget()
        self.setup_reports_tab()
        self.tabs.addTab(self.reports_tab, "📊 Отчеты")

        layout.addWidget(self.tabs)
        
        # Загружаем данные ПОСЛЕ инициализации всех UI элементов
        self.load_assets_data()
        
        # Если админ, загружаем аккаунты
        if self.current_user.get('role') == 'admin':
            self.load_accounts_data()

        print(" Интерфейс инициализирован")

    def create_menu(self):
        """Создание меню приложения"""
        menubar = self.menuBar()

        # Меню Файл
        file_menu = menubar.addMenu("📁 Файл")

        export_action = QAction("📤 Экспорт всех данных", self)
        export_action.triggered.connect(self.export_all_data)
        file_menu.addAction(export_action)

        backup_action = QAction("💾 Создать резервную копию", self)
        backup_action.triggered.connect(self.create_backup)
        file_menu.addAction(backup_action)

        file_menu.addSeparator()

        exit_action = QAction("🚪 Выход", self)
        exit_action.triggered.connect(self.logout)
        file_menu.addAction(exit_action)

        # Меню Вид
        view_menu = menubar.addMenu("👁️ Вид")
        
        # Тёмная тема
        dark_theme_action = QAction("🌙 Тёмная тема", self)
        dark_theme_action.triggered.connect(lambda: self.set_theme('dark'))
        view_menu.addAction(dark_theme_action)
        
        # Светлая тема
        light_theme_action = QAction("☀️ Светлая тема", self)
        light_theme_action.triggered.connect(lambda: self.set_theme('light'))
        view_menu.addAction(light_theme_action)

        # Меню Справка
        help_menu = menubar.addMenu("❓ Справка")

        help_action = QAction("📘 Руководство пользователя", self)
        help_action.triggered.connect(self.show_help)
        help_menu.addAction(help_action)

        about_action = QAction("ℹ️ О программе", self)
        about_action.triggered.connect(self.show_about)
        help_menu.addAction(about_action)

        update_action = QAction("🔄 Проверить обновления", self)
        update_action.triggered.connect(self.check_for_updates)
        help_menu.addAction(update_action)

    def setup_dashboard_tab(self):
        """Настройка вкладки панели управления"""
        layout = QVBoxLayout(self.dashboard_tab)

        # Заголовок
        title_label = QLabel("📊 Панель управления системой учета")
        title_label.setStyleSheet("font-size: 18px; font-weight: bold; padding: 10px;")
        layout.addWidget(title_label)

        # Статистика в виде сетки
        stats_grid = QGridLayout()

        # Виджеты статистики
        self.total_assets_label = self.create_stat_widget("Всего активов", "0")
        self.available_assets_label = self.create_stat_widget("Доступно", "0")
        self.issued_assets_label = self.create_stat_widget("Выдано", "0")
        self.overdue_assets_label = self.create_stat_widget("Просрочено", "0")
        self.employees_label = self.create_stat_widget("Сотрудников", "0")
        self.total_operations_label = self.create_stat_widget("Операций", "0")

        # Располагаем виджеты в сетке 2x3
        stats_grid.addWidget(self.total_assets_label, 0, 0)
        stats_grid.addWidget(self.available_assets_label, 0, 1)
        stats_grid.addWidget(self.issued_assets_label, 0, 2)
        stats_grid.addWidget(self.overdue_assets_label, 1, 0)
        stats_grid.addWidget(self.employees_label, 1, 1)
        stats_grid.addWidget(self.total_operations_label, 1, 2)

        layout.addLayout(stats_grid)

        # Разделитель
        separator = QFrame()
        separator.setFrameShape(QFrame.Shape.HLine)
        separator.setFrameShadow(QFrame.Shadow.Sunken)
        layout.addWidget(separator)

        # Последние операции
        recent_ops_label = QLabel("📝 Последние операции")
        recent_ops_label.setStyleSheet("font-size: 16px; font-weight: bold; padding: 10px;")
        layout.addWidget(recent_ops_label)

        self.recent_operations_table = QTableView()
        layout.addWidget(self.recent_operations_table)

        # Кнопка обновления
        refresh_btn = QPushButton("🔄 Обновить панель")
        refresh_btn.clicked.connect(self.update_dashboard)
        layout.addWidget(refresh_btn)

        # Обновляем данные при открытии вкладки
        self.tabs.currentChanged.connect(self.on_tab_changed)

    def create_stat_widget(self, title, value):
        """Создание виджета статистики"""
        widget = QWidget()
        widget.setStyleSheet("""
            QWidget {
                background-color: #f0f0f0;
                border-radius: 8px;
                padding: 10px;
                margin: 5px;
            }
            QLabel {
                font-weight: bold;
                color: #333;
            }
        """)

        layout = QVBoxLayout(widget)

        title_label = QLabel(title)
        title_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        title_label.setStyleSheet("font-size: 12px; color: #666;")

        value_label = QLabel(value)
        value_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        value_label.setStyleSheet("font-size: 24px; font-weight: bold; color: #2c3e50;")

        layout.addWidget(title_label)
        layout.addWidget(value_label)

        # Сохраняем ссылку на label с значением
        if title == "Всего активов":
            self.total_assets_value = value_label
        elif title == "Доступно":
            self.available_assets_value = value_label
        elif title == "Выдано":
            self.issued_assets_value = value_label
        elif title == "Просрочено":
            self.overdue_assets_value = value_label
        elif title == "Сотрудников":
            self.employees_value = value_label
        elif title == "Операций":
            self.total_operations_value = value_label

        return widget

    def on_tab_changed(self, index):
        """Обработчик смены вкладки"""
        tab_text = self.tabs.tabText(index)
        
        if tab_text == "🏠 Панель управления":
            self.update_dashboard()
        elif tab_text == "📋 Каталог активов":
            self.load_assets_data()
        elif tab_text == "🔄 Операции":
            self.load_history_data()

    def update_dashboard(self):
        """Обновление данных на панели управления"""
        print(" Обновление панели управления...")

        try:
            # Получаем статистику
            total_assets = self.db.execute_query("SELECT COUNT(*) FROM Assets")[0][0]
            available_assets = \
            self.db.execute_query("SELECT COUNT(*) FROM Assets WHERE current_status = 'Доступен'")[0][0]
            
            # Подсчитываем КОЛИЧЕСТВО выданных единиц (не активов, а выданных штук)
            issued_count_result = self.db.execute_query("""
                SELECT COUNT(*)
                FROM Usage_History
                WHERE operation_type = 'выдача'
                  AND actual_return_date IS NULL
            """)
            issued_assets = issued_count_result[0][0] if issued_count_result else 0

            overdue_assets = self.db.execute_query("""
                SELECT COUNT(*) 
                FROM Usage_History uh
                JOIN Assets a ON uh.asset_id = a.asset_id
                WHERE uh.operation_type = 'выдача'
                    AND uh.actual_return_date IS NULL
                    AND DATE(uh.planned_return_date) < DATE('now')
            """)[0][0]

            total_employees = self.db.execute_query("SELECT COUNT(*) FROM Employees")[0][0]
            total_operations = self.db.execute_query("SELECT COUNT(*) FROM Usage_History")[0][0]

            # Обновляем значения
            self.total_assets_value.setText(str(total_assets))
            self.available_assets_value.setText(str(available_assets))
            self.issued_assets_value.setText(str(issued_assets))
            self.overdue_assets_value.setText(str(overdue_assets))
            self.employees_value.setText(str(total_employees))
            self.total_operations_value.setText(str(total_operations))

            # Загружаем последние операции
            self.load_recent_operations()

            print(" Панель управления обновлена")

        except Exception as e:
            print(f" Ошибка обновления панели управления: {e}")

    def load_recent_operations(self):
        """Загрузка последних операций для дашборда"""
        if not hasattr(self, 'db_connection'):
            return

        # Очищаем старую модель
        if hasattr(self, 'recent_operations_table') and self.recent_operations_table.model():
            self.recent_operations_table.setModel(None)

        model = QSqlQueryModel()

        query = """
        SELECT 
            CASE 
                WHEN uh.operation_type = 'выдача' THEN '📤 Выдача'
                WHEN uh.operation_type = 'возврат' THEN '📥 Возврат'
                WHEN uh.operation_type = 'списание' THEN '🗑️ Списание'
                ELSE uh.operation_type
            END as 'Тип операции',
            a.name as 'Актив',
            e.last_name || ' ' || e.first_name as 'Сотрудник',
            COALESCE(uh.notes, '') as 'Кол-во / Примечание',
            uh.operation_date as 'Дата операции'
        FROM Usage_History uh
        LEFT JOIN Employees e ON uh.employee_id = e.employee_id
        LEFT JOIN Assets a ON uh.asset_id = a.asset_id
        ORDER BY uh.history_id DESC
        LIMIT 10
        """

        model.setQuery(query, self.db_connection)
        self.recent_operations_table.setModel(model)
        self.recent_operations_table.resizeColumnsToContents()

    def setup_assets_tab(self):
        """Настройка вкладки каталога активов"""
        layout = QVBoxLayout(self.assets_tab)

        # Панель кнопок
        buttons_layout = QHBoxLayout()

        self.btn_add = QPushButton("➕ Добавить актив")
        self.btn_edit = QPushButton("✏️ Редактировать")
        self.btn_delete = QPushButton("🗑️ Удалить")
        self.btn_import = QPushButton("📥 Импорт из Excel")
        self.btn_refresh = QPushButton("🔄 Обновить")

        buttons_layout.addWidget(self.btn_add)
        buttons_layout.addWidget(self.btn_edit)
        buttons_layout.addWidget(self.btn_delete)
        buttons_layout.addWidget(self.btn_import)
        buttons_layout.addWidget(self.btn_refresh)
        buttons_layout.addStretch()

        layout.addLayout(buttons_layout)

        # Таблица активов
        self.assets_table = QTableView()
        self.assets_table.doubleClicked.connect(self.edit_asset)
        layout.addWidget(self.assets_table)

        # Подключаем кнопки
        self.btn_refresh.clicked.connect(self.load_assets_data)
        self.btn_add.clicked.connect(self.add_asset)
        self.btn_edit.clicked.connect(self.edit_asset)
        self.btn_delete.clicked.connect(self.delete_asset)
        self.btn_import.clicked.connect(self.import_assets_from_excel)

    def setup_operations_tab(self):
        """Настройка вкладки операций"""
        layout = QVBoxLayout(self.operations_tab)

        # Панель фильтров истории
        filter_layout = QHBoxLayout()

        self.history_employee_filter = QComboBox()
        self.history_employee_filter.addItem("Все сотрудники", None)

        self.history_operation_filter = QComboBox()
        self.history_operation_filter.addItem("Все операции", None)
        self.history_operation_filter.addItem("Выдача", "выдача")
        self.history_operation_filter.addItem("Возврат", "возврат")
        self.history_operation_filter.addItem("Списание", "списание")

        self.history_date_from = QDateEdit()
        self.history_date_from.setDate(QDate.currentDate().addDays(-30))
        self.history_date_from.setCalendarPopup(True)

        self.history_date_to = QDateEdit()
        self.history_date_to.setDate(QDate.currentDate())
        self.history_date_to.setCalendarPopup(True)

        self.btn_apply_filters = QPushButton("🔍 Применить фильтры")
        self.btn_clear_filters = QPushButton("❌ Сбросить")

        filter_layout.addWidget(QLabel("Сотрудник:"))
        filter_layout.addWidget(self.history_employee_filter)
        filter_layout.addWidget(QLabel("Операция:"))
        filter_layout.addWidget(self.history_operation_filter)
        filter_layout.addWidget(QLabel("С:"))
        filter_layout.addWidget(self.history_date_from)
        filter_layout.addWidget(QLabel("По:"))
        filter_layout.addWidget(self.history_date_to)
        filter_layout.addWidget(self.btn_apply_filters)
        filter_layout.addWidget(self.btn_clear_filters)
        filter_layout.addStretch()

        layout.addLayout(filter_layout)

        # Панель кнопок операций
        operations_layout = QHBoxLayout()

        self.btn_issue = QPushButton("📤 Выдать актив")
        self.btn_return = QPushButton("📥 Вернуть актив")
        self.btn_request = QPushButton("📝 Запросить актив")  # Новая кнопка для пользователей
        self.btn_history = QPushButton("🔄 Обновить историю")

        operations_layout.addWidget(self.btn_issue)
        operations_layout.addWidget(self.btn_return)
        operations_layout.addWidget(self.btn_request)
        operations_layout.addWidget(self.btn_history)
        
        # Скрываем кнопку выдачи для обычных пользователей
        if self.current_user.get('role') != 'admin':
            self.btn_issue.hide()
        
        operations_layout.addStretch()

        layout.addLayout(operations_layout)

        # Таблица для истории операций
        self.history_table = QTableView()
        layout.addWidget(self.history_table)

        # Подключаем кнопки
        self.btn_issue.clicked.connect(self.issue_asset)
        self.btn_return.clicked.connect(self.return_asset)
        self.btn_request.clicked.connect(self.request_asset)
        self.btn_history.clicked.connect(self.load_history_data)
        self.btn_apply_filters.clicked.connect(self.load_history_data)
        self.btn_clear_filters.clicked.connect(self.clear_history_filters)

        # Загружаем данные для фильтров
        self.load_history_filters_data()

    def setup_reports_tab(self):
        """Настройка вкладки отчетов"""
        layout = QVBoxLayout(self.reports_tab)

        # Панель кнопок отчетов
        reports_buttons_layout = QHBoxLayout()

        self.btn_overdue_report = QPushButton("📅 Отчет по просрочкам")
        self.btn_usage_report = QPushButton("📈 Отчет по использованию")
        self.btn_inventory_report = QPushButton("📋 Инвентаризационная ведомость")

        # Создаем группу кнопок для эксклюзивного выбора
        self.reports_button_group = QButtonGroup()
        self.reports_button_group.addButton(self.btn_overdue_report, 0)
        self.reports_button_group.addButton(self.btn_usage_report, 1)
        self.reports_button_group.addButton(self.btn_inventory_report, 2)
        
        # Делаем кнопки переключаемыми (checkable)
        for button in [self.btn_overdue_report, self.btn_usage_report, self.btn_inventory_report]:
            button.setCheckable(True)
        
        # По умолчанию выбираем первую кнопку
        self.btn_overdue_report.setChecked(True)

        reports_buttons_layout.addWidget(self.btn_overdue_report)
        reports_buttons_layout.addWidget(self.btn_usage_report)
        reports_buttons_layout.addWidget(self.btn_inventory_report)
        reports_buttons_layout.addStretch()

        layout.addLayout(reports_buttons_layout)

        # Панель кнопок экспорта
        export_buttons_layout = QHBoxLayout()

        self.btn_export_csv = QPushButton("💾 Экспорт в CSV")
        self.btn_export_excel = QPushButton("📊 Экспорт в Excel")

        export_buttons_layout.addWidget(self.btn_export_csv)
        export_buttons_layout.addWidget(self.btn_export_excel)
        export_buttons_layout.addStretch()

        layout.addLayout(export_buttons_layout)

        # Таблица для отображения отчетов
        self.reports_table = QTableView()
        layout.addWidget(self.reports_table)

        # Подключаем кнопки
        self.btn_overdue_report.clicked.connect(self.generate_overdue_report)
        self.btn_usage_report.clicked.connect(self.generate_usage_report)
        self.btn_inventory_report.clicked.connect(self.generate_inventory_report)
        self.btn_export_csv.clicked.connect(self.export_to_csv)
        self.btn_export_excel.clicked.connect(self.export_to_excel)

        # Текущий тип отчета для экспорта
        self.current_report_type = None

    def load_assets_data(self):
        """Загрузка данных об активах"""
        print(" Загрузка данных об активах...")

        if not hasattr(self, 'db_connection'):
            self.db_connection = QSqlDatabase.addDatabase("QSQLITE")
            self.db_connection.setDatabaseName("inventory.db")

        if not self.db_connection.isOpen():
            if not self.db_connection.open():
                error = self.db_connection.lastError().text()
                print(f" Ошибка подключения к базе: {error}")
                QMessageBox.critical(self, "Ошибка", f"Не удалось подключиться к базе данных!\n{error}")
                return

        model = QSqlQueryModel()

        query = """
        SELECT 
            a.asset_id as 'ID',
            a.name as 'Название',
            at.type_name as 'Тип',
            a.model as 'Модель',
            a.serial_number as 'Серийный номер',
            a.current_status as 'Статус',
            l.location_name as 'Местоположение',
            a.quantity as 'Количество'
        FROM Assets a
        JOIN Asset_Types at ON a.type_id = at.type_id
        JOIN Locations l ON a.location_id = l.location_id
        ORDER BY a.asset_id
        """

        model.setQuery(query)

        if model.lastError().isValid():
            error = model.lastError().text()
            print(f" Ошибка выполнения запроса: {error}")
            QMessageBox.critical(self, "Ошибка", f"Ошибка загрузки данных:\n{error}")
        else:
            row_count = model.rowCount()
            print(f" Данные загружены. Найдено записей: {row_count}")

            if row_count == 0:
                print("️ В базе данных нет записей.")

        # Устанавливаем модель только если таблица существует (для админов)
        if hasattr(self, 'assets_table'):
            self.assets_table.setModel(model)
            self.assets_table.resizeColumnsToContents()

    def load_history_data(self):
        """Загрузка истории операций с фильтрами"""
        print(" Загрузка истории операций...")

        # Сначала обновляем статусы просроченных активов в примечаниях
        self._update_overdue_notes()

        if not hasattr(self, 'db_connection'):
            return

        model = QSqlQueryModel()

        # Базовый запрос
        query = """
        SELECT 
            uh.history_id as 'ID',
            e.last_name || ' ' || e.first_name as 'Сотрудник',
            a.name as 'Актив',
            a.model as 'Модель',
            CASE 
                WHEN uh.operation_type = 'выдача' THEN '📤 Выдача'
                WHEN uh.operation_type = 'возврат' THEN '📥 Возврат'
                WHEN uh.operation_type = 'списание' THEN '🗑️ Списание'
                ELSE uh.operation_type
            END as 'Тип операции',
            uh.operation_date as 'Дата операции',
            uh.planned_return_date as 'План возврата',
            uh.actual_return_date as 'Факт возврата',
            CASE 
                WHEN a.current_status = 'Выдан' THEN '📤 Выдан'
                WHEN a.current_status = 'Доступен' THEN '✅ Доступен'
                WHEN a.current_status = 'Списан' THEN '🗑️ Списан'
                ELSE a.current_status
            END as 'Статус актива',
            CASE 
                WHEN uh.operation_type = 'выдача' AND uh.actual_return_date IS NULL AND DATE(uh.planned_return_date) < DATE('now')
                THEN COALESCE(uh.notes, '') || ' [Просрочено]'
                WHEN uh.actual_return_date IS NOT NULL AND DATE(uh.actual_return_date) > DATE(uh.planned_return_date)
                THEN COALESCE(uh.notes, '') || ' [Возвращено с опозданием]'
                ELSE COALESCE(uh.notes, '')
            END as 'Примечания'
        FROM Usage_History uh
        JOIN Employees e ON uh.employee_id = e.employee_id
        JOIN Assets a ON uh.asset_id = a.asset_id
        WHERE 1=1
        """

        params = []

        # Применяем фильтры
        employee_filter = self.history_employee_filter.currentData()
        if employee_filter is not None:
            query += " AND uh.employee_id = ?"
            params.append(employee_filter)

        operation_filter = self.history_operation_filter.currentData()
        if operation_filter is not None:
            query += " AND uh.operation_type = ?"
            params.append(operation_filter)

        date_from = self.history_date_from.date().toString("yyyy-MM-dd")
        date_to = self.history_date_to.date().toString("yyyy-MM-dd")
        query += " AND DATE(uh.operation_date) BETWEEN ? AND ?"
        params.extend([date_from, date_to])

        query += " ORDER BY uh.operation_date DESC"

        # Выполняем запрос
        query_obj = QSqlQuery(self.db_connection)
        query_obj.prepare(query)

        for param in params:
            query_obj.addBindValue(param)

        # Очищаем старую модель
        if hasattr(self, 'history_table') and self.history_table.model():
            self.history_table.setModel(None)

        if not query_obj.exec():
            error = query_obj.lastError().text()
            print(f" Ошибка загрузки истории: {error}")
        else:
            model.setQuery(query_obj)
            print(f" История загружена. Записей: {model.rowCount()}")

        self.history_table.setModel(model)
        self.history_table.resizeColumnsToContents()

    def _update_overdue_notes(self):
        """Автоматическое обновление примечаний для просроченных активов"""
        try:
            # Получаем просроченные активы без отметки в примечаниях
            query = """
                SELECT history_id, notes
                FROM Usage_History
                WHERE operation_type = 'выдача'
                    AND actual_return_date IS NULL
                    AND DATE(planned_return_date) < DATE('now')
                    AND (notes IS NULL OR notes NOT LIKE '%Просрочено%')
            """
            
            results = self.db.execute_query(query)
            
            for history_id, notes in results:
                from PyQt6.QtCore import QDate
                new_notes = f"{notes}\n[Просрочено: {QDate.currentDate().toString('yyyy-MM-dd')}]" if notes else f"[Просрочено: {QDate.currentDate().toString('yyyy-MM-dd')}]"
                
                self.db.execute_update(
                    "UPDATE Usage_History SET notes = ? WHERE history_id = ?",
                    (new_notes, history_id)
                )
                
        except Exception as e:
            print(f"️ Предупреждение при обновлении примечаний: {e}")


    def load_history_filters_data(self):
        """Загрузка данных для фильтров истории"""
        try:
            employees = self.db.execute_query("""
                SELECT employee_id, last_name || ' ' || first_name as full_name
                FROM Employees 
                ORDER BY last_name, first_name
            """)

            for employee_id, full_name in employees:
                self.history_employee_filter.addItem(full_name, employee_id)

        except Exception as e:
            print(f"Ошибка загрузки фильтров истории: {e}")

    def clear_history_filters(self):
        """Сброс фильтров истории"""
        self.history_employee_filter.setCurrentIndex(0)
        self.history_operation_filter.setCurrentIndex(0)
        self.history_date_from.setDate(QDate.currentDate().addDays(-30))
        self.history_date_to.setDate(QDate.currentDate())
        self.load_history_data()

    def get_selected_asset_id(self):
        """Получение ID выбранного актива"""
        # Получаем текущую выбранную строку
        current_index = self.assets_table.currentIndex()
        if not current_index.isValid():
            QMessageBox.warning(self, "Ошибка", "Выберите актив из таблицы!")
            return None

        # Получаем модель и данные из первой колонки (ID)
        model = self.assets_table.model()
        row = current_index.row()
        index_id = model.index(row, 0)  # Первая колонка - ID
        asset_id = model.data(index_id)

        return asset_id

    def add_asset(self):
        """Добавление нового актива"""
        print("➕ Открытие диалога добавления актива...")
        dialog = AssetDialog(self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self._refresh_all_data()

    def edit_asset(self):
        """Редактирование выбранного актива"""
        print(" Попытка редактирования актива...")
        asset_id = self.get_selected_asset_id()
        if asset_id is None:
            return

        print(f"Редактирование актива ID: {asset_id}")
        dialog = EditAssetDialog(asset_id, self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            print(" Актив успешно отредактирован")
            self._refresh_all_data()

    def delete_asset(self):
        """Удаление выбранного актива"""
        print("Попытка удаления актива...")
        asset_id = self.get_selected_asset_id()
        if asset_id is None:
            return

        print(f"Удаление актива ID: {asset_id}")

        # Создаем диалог для удаления
        dialog = EditAssetDialog(asset_id, self)

        # Подключаемся к сигналу закрытия диалога
        dialog.finished.connect(lambda result: self.on_asset_dialog_finished(result, asset_id))

        # Показываем диалог - пользователь сам нажмет кнопку удаления внутри диалога
        dialog.exec()

    def _refresh_all_data(self):
        """Обновление всех таблиц и панелей после операции"""
        self.load_assets_data()
        self.load_history_data()
        self.load_recent_operations()
        
        # Обновляем статистику на панели управления
        if hasattr(self, 'total_assets_value'):
            self.update_dashboard()

    def on_asset_dialog_finished(self, result, asset_id):
        """Обработчик завершения работы диалога редактирования/удаления"""
        if result == QDialog.DialogCode.Accepted:
            print(" Операция с активом завершена, обновляем таблицы...")
            self._refresh_all_data()

    def issue_asset(self):
        """Выдача актива сотруднику"""
        print("Открытие диалога выдачи актива...")
        dialog = IssueDialog(self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self._refresh_all_data()

    def return_asset(self):
        """Возврат актива"""
        print("Открытие диалога возврата актива...")
        dialog = ReturnDialog(self, self.current_user)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self._refresh_all_data()

    def export_to_csv(self):
        """Экспорт текущего отчета в CSV"""
        if not self.reports_table.model():
            QMessageBox.warning(self, "Ошибка", "Сначала сгенерируйте отчет!")
            return

        # Выбираем путь для сохранения файла
        report_name = self.current_report_type or "report"
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Сохранить отчет как CSV",
            f"{report_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
            "CSV Files (*.csv);;All Files (*)"
        )

        if not file_path:
            return

        try:
            model = self.reports_table.model()
            
            with open(file_path, 'w', newline='', encoding='utf-8-sig') as csvfile:
                # Получаем количество столбцов и строк
                row_count = model.rowCount()
                col_count = model.columnCount()
                
                # Пишем заголовки
                headers = []
                for col in range(col_count):
                    header = model.headerData(col, Qt.Orientation.Horizontal)
                    headers.append(str(header) if header else "")
                
                writer = csv.writer(csvfile, delimiter=';')
                writer.writerow(headers)
                
                # Пишем данные
                for row in range(row_count):
                    row_data = []
                    for col in range(col_count):
                        index = model.index(row, col)
                        value = model.data(index)
                        row_data.append(str(value) if value is not None else "")
                    writer.writerow(row_data)
            
            QMessageBox.information(self, "Успех", f"Отчет успешно сохранен:\n{file_path}")
            
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка экспорта в CSV: {e}")

    def export_to_excel(self):
        """Экспорт текущего отчета в Excel"""
        if not self.reports_table.model():
            QMessageBox.warning(self, "Ошибка", "Сначала сгенерируйте отчет!")
            return

        # Выбираем путь для сохранения файла
        report_name = self.current_report_type or "report"
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Сохранить отчет как Excel",
            f"{report_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "Excel Files (*.xlsx);;All Files (*)"
        )

        if not file_path:
            return

        try:
            model = self.reports_table.model()
            wb = Workbook()
            ws = wb.active
            ws.title = "Отчет"
            
            # Получаем количество столбцов и строк
            row_count = model.rowCount()
            col_count = model.columnCount()
            
            # Стили для заголовка
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            # Пишем заголовки
            for col in range(col_count):
                header = model.headerData(col, Qt.Orientation.Horizontal)
                header_text = str(header) if header else ""
                cell = ws.cell(row=1, column=col+1, value=header_text)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Пишем данные
            for row in range(row_count):
                for col in range(col_count):
                    index = model.index(row, col)
                    value = model.data(index)
                    cell = ws.cell(row=row+2, column=col+1, value=value)
                    cell.alignment = Alignment(wrap_text=True)
            
            # Автоматически регулируем ширину столбцов
            for col in range(col_count):
                max_length = 0
                column_letter = chr(65 + col) if col < 26 else "A" + chr(65 + col - 26)
                
                for row in range(row_count + 1):
                    try:
                        cell_value = str(ws.cell(row=row+1, column=col+1).value or "")
                        if len(cell_value) > max_length:
                            max_length = len(cell_value)
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[chr(65 + col) if col < 26 else "A" + chr(65 + col - 26)].width = adjusted_width
            
            wb.save(file_path)
            QMessageBox.information(self, "Успех", f"Отчет успешно сохранен:\n{file_path}")
            
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка экспорта в Excel: {e}")

    def print_report(self):
        """Метод больше не используется"""
        pass

    def generate_overdue_report(self):
        """Генерация отчета по просроченным активам"""
        print("Генерация отчета по просрочкам...")
        self.current_report_type = "overdue_report"

        if not hasattr(self, 'db_connection'):
            return

        # Очищаем старую модель
        if hasattr(self, 'reports_table') and self.reports_table.model():
            self.reports_table.setModel(None)

        model = QSqlQueryModel()

        query = """
        SELECT 
            a.name as 'Актив',
            a.model as 'Модель',
            e.last_name || ' ' || e.first_name as 'Сотрудник',
            uh.operation_date as 'Дата выдачи',
            uh.planned_return_date as 'Плановый возврат',
            uh.actual_return_date as 'Фактический возврат',
            CASE 
                WHEN uh.actual_return_date IS NULL 
                THEN CAST(JULIANDAY('now') - JULIANDAY(uh.planned_return_date) AS INTEGER)
                ELSE CAST(JULIANDAY(uh.actual_return_date) - JULIANDAY(uh.planned_return_date) AS INTEGER)
            END as 'Дней просрочки',
            CASE 
                WHEN uh.actual_return_date IS NULL THEN '⏰ Ещё не возвращен'
                WHEN uh.actual_return_date IS NOT NULL AND DATE(uh.actual_return_date) > DATE(uh.planned_return_date)
                THEN '⚠️ Возвращено с опозданием'
                ELSE ''
            END as 'Статус'
        FROM Usage_History uh
        JOIN Assets a ON uh.asset_id = a.asset_id
        JOIN Employees e ON uh.employee_id = e.employee_id
        WHERE uh.operation_type = 'выдача'
            AND (
                (uh.actual_return_date IS NULL AND DATE(uh.planned_return_date) < DATE('now'))
                OR
                (uh.actual_return_date IS NOT NULL AND DATE(uh.actual_return_date) > DATE(uh.planned_return_date))
            )
        ORDER BY uh.planned_return_date
        """

        model.setQuery(query, self.db_connection)
        self.reports_table.setModel(model)
        self.reports_table.resizeColumnsToContents()

        if model.rowCount() == 0:
            QMessageBox.information(self, "Информация", "Нет просроченных активов!")

    def generate_usage_report(self):
        """Генерация отчета по использованию активов"""
        print("Генерация отчета по использованию...")
        self.current_report_type = "usage_report"

        if not hasattr(self, 'db_connection'):
            return

        model = QSqlQueryModel()

        query = """
        SELECT 
            a.name as 'Актив',
            a.model as 'Модель',
            COUNT(uh.history_id) as 'Количество выдач',
            MIN(uh.operation_date) as 'Первая выдача',
            MAX(uh.operation_date) as 'Последняя выдача'
        FROM Assets a
        LEFT JOIN Usage_History uh ON a.asset_id = uh.asset_id AND uh.operation_type = 'выдача'
        GROUP BY a.asset_id
        ORDER BY COUNT(uh.history_id) DESC
        """

        model.setQuery(query)
        self.reports_table.setModel(model)
        self.reports_table.resizeColumnsToContents()

    def generate_inventory_report(self):
        """Генерация инвентаризационной ведомости"""
        print("Генерация инвентаризационной ведомости...")
        self.current_report_type = "inventory_report"

        if not hasattr(self, 'db_connection'):
            return

        model = QSqlQueryModel()

        query = """
        SELECT 
            a.asset_id as 'Инвентарный номер',
            a.name as 'Наименование',
            at.type_name as 'Тип',
            a.model as 'Модель',
            a.serial_number as 'Серийный номер',
            a.current_status as 'Статус',
            l.location_name as 'Местоположение',
            a.quantity as 'Количество'
        FROM Assets a
        JOIN Asset_Types at ON a.type_id = at.type_id
        JOIN Locations l ON a.location_id = l.location_id
        ORDER BY a.asset_id
        """

        model.setQuery(query)
        self.reports_table.setModel(model)
        self.reports_table.resizeColumnsToContents()

    def export_all_data(self):
        """Экспорт всех данных системы в Excel"""
        # Выбираем путь для сохранения файла
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Экспорт всех данных",
            f"export_all_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "Excel Files (*.xlsx);;All Files (*)"
        )

        if not file_path:
            return

        try:
            wb = Workbook()
            wb.remove(wb.active)  # Удаляем лист по умолчанию
            
            # Стили для заголовков
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            data_alignment = Alignment(wrap_text=True, vertical="top")
            
            # 1. Лист с активами
            self._export_assets_sheet(wb, header_font, header_fill, header_alignment, data_alignment)
            
            # 2. Лист с сотрудниками
            self._export_employees_sheet(wb, header_font, header_fill, header_alignment, data_alignment)
            
            # 3. Лист с историей операций
            self._export_history_sheet(wb, header_font, header_fill, header_alignment, data_alignment)
            
            # 4. Лист с типами активов
            self._export_asset_types_sheet(wb, header_font, header_fill, header_alignment, data_alignment)
            
            # 5. Лист с местоположениями
            self._export_locations_sheet(wb, header_font, header_fill, header_alignment, data_alignment)
            
            # 6. Лист со статистикой
            self._export_statistics_sheet(wb, header_font, header_fill, header_alignment, data_alignment)
            
            wb.save(file_path)
            QMessageBox.information(self, "Успех", f"Все данные успешно экспортированы:\n{file_path}")
            
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка при экспорте данных: {e}")

    def _export_assets_sheet(self, wb, header_font, header_fill, header_alignment, data_alignment):
        """Экспорт таблицы активов"""
        ws = wb.create_sheet("Активы")
        
        headers = ["ID", "Название", "Тип", "Модель", "Серийный номер", "Статус", "Местоположение", "Количество"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Получаем данные активов
        query = """
        SELECT a.asset_id, a.name, at.type_name, a.model, a.serial_number, 
               a.current_status, l.location_name, a.quantity
        FROM Assets a
        JOIN Asset_Types at ON a.type_id = at.type_id
        JOIN Locations l ON a.location_id = l.location_id
        ORDER BY a.asset_id
        """
        
        rows = self.db.execute_query(query)
        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = data_alignment
        
        # Регулируем ширину столбцов
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col)].width = 20

    def _export_employees_sheet(self, wb, header_font, header_fill, header_alignment, data_alignment):
        """Экспорт таблицы сотрудников"""
        ws = wb.create_sheet("Сотрудники")
        
        headers = ["ID", "Фамилия", "Имя", "Отчество", "Должность", "Email", "Телефон"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Получаем данные сотрудников с присоединением таблицы должностей
        query = """
        SELECT e.employee_id, e.last_name, e.first_name, e.patronymic, 
               COALESCE(p.position_name, ''), e.email, e.phone 
        FROM Employees e
        LEFT JOIN Positions p ON e.position_id = p.position_id
        ORDER BY e.employee_id
        """
        
        rows = self.db.execute_query(query)
        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = data_alignment
        
        # Регулируем ширину столбцов
        widths = [10, 15, 15, 15, 20, 25, 15]
        for col, width in enumerate(widths, 1):
            ws.column_dimensions[chr(64 + col)].width = width

    def _export_history_sheet(self, wb, header_font, header_fill, header_alignment, data_alignment):
        """Экспорт таблицы истории операций"""
        ws = wb.create_sheet("История операций")
        
        headers = ["ID", "Актив", "Сотрудник", "Тип операции", "Дата операции", 
                   "Плановый возврат", "Фактический возврат", "Примечания"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Получаем данные истории
        query = """
        SELECT uh.history_id, a.name, e.last_name || ' ' || e.first_name, 
               uh.operation_type, uh.operation_date, uh.planned_return_date, 
               uh.actual_return_date, COALESCE(uh.notes, '')
        FROM Usage_History uh
        JOIN Assets a ON uh.asset_id = a.asset_id
        JOIN Employees e ON uh.employee_id = e.employee_id
        ORDER BY uh.history_id DESC
        """
        
        rows = self.db.execute_query(query)
        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = data_alignment
        
        # Регулируем ширину столбцов
        widths = [10, 25, 25, 15, 20, 20, 20, 40]
        for col, width in enumerate(widths, 1):
            ws.column_dimensions[chr(64 + col)].width = width

    def _export_asset_types_sheet(self, wb, header_font, header_fill, header_alignment, data_alignment):
        """Экспорт таблицы типов активов"""
        ws = wb.create_sheet("Типы активов")
        
        headers = ["ID", "Название типа"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Получаем данные типов
        query = "SELECT type_id, type_name FROM Asset_Types ORDER BY type_id"
        
        rows = self.db.execute_query(query)
        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = data_alignment
        
        # Регулируем ширину столбцов
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 30

    def _export_locations_sheet(self, wb, header_font, header_fill, header_alignment, data_alignment):
        """Экспорт таблицы местоположений"""
        ws = wb.create_sheet("Местоположения")
        
        headers = ["ID", "Название местоположения"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Получаем данные местоположений
        query = "SELECT location_id, location_name FROM Locations ORDER BY location_id"
        
        rows = self.db.execute_query(query)
        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = data_alignment
        
        # Регулируем ширину столбцов
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 40

    def _export_statistics_sheet(self, wb, header_font, header_fill, header_alignment, data_alignment):
        """Экспорт листа со статистикой"""
        ws = wb.create_sheet("Статистика", 0)  # Добавляем в начало
        
        ws.title = "Статистика"
        
        # Заголовок
        title_cell = ws.cell(row=1, column=1, value="СТАТИСТИКА СИСТЕМЫ")
        title_cell.font = Font(bold=True, size=14, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells('A1:B1')
        ws.row_dimensions[1].height = 25
        
        # Пустая строка
        ws.row_dimensions[2].height = 5
        
        # Получаем статистику
        row = 3
        
        # Всего активов
        total_assets = self.db.execute_query("SELECT COUNT(*) FROM Assets")[0][0]
        ws.cell(row=row, column=1, value="Всего активов:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=total_assets)
        row += 1
        
        # Доступны
        available = self.db.execute_query("SELECT COUNT(*) FROM Assets WHERE current_status = 'Доступен'")[0][0]
        ws.cell(row=row, column=1, value="Доступно активов:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=available)
        row += 1
        
        # Выданы
        issued = self.db.execute_query("""
            SELECT COUNT(*) FROM Usage_History 
            WHERE operation_type = 'выдача' AND actual_return_date IS NULL
        """)[0][0]
        ws.cell(row=row, column=1, value="Выдано активов:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=issued)
        row += 1
        
        # Списаны
        written_off = self.db.execute_query("SELECT COUNT(*) FROM Assets WHERE current_status = 'Списан'")[0][0]
        ws.cell(row=row, column=1, value="Списано активов:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=written_off)
        row += 1
        
        # Просроченные
        overdue = self.db.execute_query("""
            SELECT COUNT(*) FROM Usage_History uh
            WHERE uh.operation_type = 'выдача'
                AND uh.actual_return_date IS NULL
                AND DATE(uh.planned_return_date) < DATE('now')
        """)[0][0]
        ws.cell(row=row, column=1, value="Просроченные активы:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=overdue)
        row += 2
        
        # Количество сотрудников
        employees = self.db.execute_query("SELECT COUNT(*) FROM Employees")[0][0]
        ws.cell(row=row, column=1, value="Количество сотрудников:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=employees)
        row += 1
        
        # Всего операций
        operations = self.db.execute_query("SELECT COUNT(*) FROM Usage_History")[0][0]
        ws.cell(row=row, column=1, value="Всего операций:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=operations)
        row += 2
        
        # Дата экспорта
        ws.cell(row=row, column=1, value="Дата экспорта:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        
        # Регулируем ширину столбцов
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20

    def create_backup(self):
        """Создание резервной копии базы данных"""
        QMessageBox.information(self, "Резервное копирование",
                                "Функция резервного копирования будет реализована в следующей версии")

    def show_about(self):
        """Показ информации о программе"""
        about_text = """
        <h2>Система учета инструментов и расходников</h2>
        <p><b>Версия:</b> 1.0.0</p>
        <p><b>Разработчик:</b> РТУ МИРЭА</p>
        <p><b>Заказчик:</b> АО "КОНСИСТ-ОС"</p>
        <p><b>Год разработки:</b> 2025</p>

        <h3>Возможности системы:</h3>
        <ul>
            <li>Учет инструментов и расходных материалов</li>
            <li>Выдача и возврат активов сотрудникам</li>
            <li>Контроль сроков использования</li>
            <li>Формирование отчетов и аналитика</li>
            <li>Экспорт данных в различные форматы</li>
        </ul>

        <p>Система разработана в соответствии с требованиями:</p>
        <ul>
            <li>ГОСТ Р 59793-2021</li>
            <li>ГОСТ 34.602-2020</li>
            <li>Методические указания Госкорпорации "Росатом"</li>
        </ul>
        """

        QMessageBox.about(self, "О программе", about_text)

    def show_help(self):
        """Показ руководства пользователя"""
        help_text = """
        <h2>Руководство пользователя</h2>

        <h3>1. Панель управления</h3>
        <p>На главной вкладке отображается общая статистика системы.</p>

        <h3>2. Каталог активов</h3>
        <p>Работа с инструментами и расходниками:</p>
        <ul>
            <li><b>Добавить актив:</b> создание новой записи об инструменте</li>
            <li><b>Редактировать:</b> изменение данных выбранного актива</li>
            <li><b>Удалить:</b> удаление актива из системы (только доступные)</li>
            <li><b>Фильтры:</b> поиск и сортировка активов по различным параметрам</li>
        </ul>

        <h3>3. Операции</h3>
        <p>Управление выдачей и возвратом:</p>
        <ul>
            <li><b>Выдать актив:</b> оформление выдачи инструмента сотруднику</li>
            <li><b>Вернуть актив:</b> оформление возврата на склад</li>
            <li><b>История операций:</b> просмотр всех транзакций с фильтрами</li>
        </ul>

        <h3>4. Отчеты</h3>
        <p>Аналитика и документооборот:</p>
        <ul>
            <li><b>Отчет по просрочкам:</b> список активов с истекшим сроком возврата</li>
            <li><b>Отчет по использованию:</b> статистика частоты использования</li>
            <li><b>Инвентаризационная ведомость:</b> полный перечень активов</li>
            <li><b>Экспорт:</b> сохранение отчетов в CSV или Excel формате</li>
        </ul>

        <h3>5. Быстрые клавиши</h3>
        <ul>
            <li><b>F5:</b> Обновить текущую вкладку</li>
            <li><b>Ctrl+N:</b> Добавить новый актив</li>
            <li><b>Ctrl+E:</b> Редактировать выбранный актив</li>
            <li><b>Ctrl+F:</b> Поиск в текущей таблице</li>
            <li><b>Ctrl+S:</b> Сохранить/экспортировать отчет</li>
        </ul>

        <p>Для получения дополнительной помощи обратитесь к администратору системы.</p>
        """

        dialog = QDialog(self)
        dialog.setWindowTitle("Руководство пользователя")
        dialog.setFixedSize(600, 700)

        layout = QVBoxLayout(dialog)

        text_edit = QTextEdit()
        text_edit.setReadOnly(True)
        text_edit.setHtml(help_text)

        layout.addWidget(text_edit)

        close_btn = QPushButton("Закрыть")
        close_btn.clicked.connect(dialog.accept)
        layout.addWidget(close_btn)

        dialog.exec()

    def check_for_updates(self):
        """Проверка обновлений"""
        QMessageBox.information(self, "Обновления", "Проверка обновлений...\nУ вас установлена последняя версия.")

    def import_assets_from_excel(self):
        """Импорт активов из Excel файла"""
        print("Открытие диалога импорта активов из Excel...")
        
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Выбрать файл для импорта",
            "",
            "Excel Files (*.xlsx);;All Files (*)"
        )

        if not file_path:
            return

        try:
            wb = load_workbook(file_path)
            ws = wb.active

            # Пропускаем заголовок и читаем данные
            assets_count = 0
            errors = []

            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # Распаковываем данные: название, тип, модель, серийный номер, местоположение, количество
                    if len(row) < 6:
                        errors.append(f"Строка {row_idx}: недостаточно данных")
                        continue

                    name = str(row[0]).strip() if row[0] else None
                    type_name = str(row[1]).strip() if row[1] else None
                    model = str(row[2]).strip() if row[2] else ""
                    serial_number = str(row[3]).strip() if row[3] else ""
                    location_name = str(row[4]).strip() if row[4] else None
                    quantity = int(row[5]) if row[5] and str(row[5]).isdigit() else 1

                    # Проверяем обязательные поля
                    if not name:
                        errors.append(f"Строка {row_idx}: отсутствует название актива")
                        continue
                    if not type_name:
                        errors.append(f"Строка {row_idx}: отсутствует тип актива")
                        continue
                    if not location_name:
                        errors.append(f"Строка {row_idx}: отсутствует местоположение")
                        continue

                    # Получаем или создаем тип актива
                    type_id = self._get_or_create_asset_type(type_name)
                    if not type_id:
                        errors.append(f"Строка {row_idx}: не удалось создать тип '{type_name}'")
                        continue

                    # Получаем или создаем местоположение
                    location_id = self._get_or_create_location(location_name)
                    if not location_id:
                        errors.append(f"Строка {row_idx}: не удалось создать местоположение '{location_name}'")
                        continue

                    # Добавляем актив в базу данных
                    query = """
                    INSERT INTO Assets (name, type_id, model, serial_number, location_id, current_status, quantity)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                    """
                    self.db.execute_update(
                        query,
                        (name, type_id, model, serial_number, location_id, "Доступен", quantity)
                    )
                    assets_count += 1
                    print(f" Актив добавлен (строка {row_idx}): {name}")

                except Exception as e:
                    errors.append(f"Строка {row_idx}: {str(e)}")
                    continue

            # Показываем результаты
            message = f"✅ Успешно импортировано активов: {assets_count}"
            if errors:
                message += f"\n\n⚠️ Ошибки при импорте ({len(errors)}):\n"
                message += "\n".join(errors[:10])  # Показываем первые 10 ошибок
                if len(errors) > 10:
                    message += f"\n... и еще {len(errors) - 10} ошибок"

            QMessageBox.information(self, "Результаты импорта", message)

            # Обновляем таблицу активов
            if assets_count > 0:
                self.load_assets_data()

        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка при импорте файла:\n{str(e)}")

    def _get_or_create_asset_type(self, type_name):
        """Получает ID типа актива, если существует, или создает новый"""
        try:
            # Проверяем, существует ли тип
            result = self.db.execute_query(
                "SELECT type_id FROM Asset_Types WHERE type_name = ?",
                (type_name,)
            )
            if result:
                return result[0][0]

            # Создаем новый тип
            new_id = self.db.execute_update(
                "INSERT INTO Asset_Types (type_name) VALUES (?)",
                (type_name,)
            )
            return new_id
        except Exception as e:
            print(f" Ошибка при работе с типом актива: {e}")
            return None

    def _get_or_create_location(self, location_name):
        """Получает ID местоположения, если существует, или создает новое"""
        try:
            # Проверяем, существует ли местоположение
            result = self.db.execute_query(
                "SELECT location_id FROM Locations WHERE location_name = ?",
                (location_name,)
            )
            if result:
                return result[0][0]

            # Создаем новое местоположение
            new_id = self.db.execute_update(
                "INSERT INTO Locations (location_name) VALUES (?)",
                (location_name,)
            )
            return new_id
        except Exception as e:
            print(f" Ошибка при работе с местоположением: {e}")
            return None

    def setup_requests_tab(self):
        """Настройка вкладки запросов на выдачу активов (только для админа)"""
        layout = QVBoxLayout(self.requests_tab)

        # Заголовок
        title_label = QLabel("📬 Запросы на выдачу активов")
        title_label.setStyleSheet("font-size: 14px; font-weight: bold; padding: 10px;")
        layout.addWidget(title_label)

        # Панель кнопок
        buttons_layout = QHBoxLayout()

        self.btn_approve_request = QPushButton("✅ Одобрить")
        self.btn_reject_request = QPushButton("❌ Отклонить")
        self.btn_refresh_requests = QPushButton("🔄 Обновить")

        buttons_layout.addWidget(self.btn_approve_request)
        buttons_layout.addWidget(self.btn_reject_request)
        buttons_layout.addWidget(self.btn_refresh_requests)
        buttons_layout.addStretch()

        layout.addLayout(buttons_layout)

        # Таблица запросов
        self.requests_table = QTableView()
        layout.addWidget(self.requests_table)

        # Подключаем кнопки
        self.btn_approve_request.clicked.connect(self.approve_request)
        self.btn_reject_request.clicked.connect(self.reject_request)
        self.btn_refresh_requests.clicked.connect(self.load_requests_data)

        # Загружаем данные
        self.load_requests_data()

    def setup_accounts_tab(self):
        """Настройка вкладки управления аккаунтами (только для админа)"""
        layout = QVBoxLayout(self.accounts_tab)
        layout.setSpacing(10)
        layout.setContentsMargins(10, 10, 10, 10)

        # Заголовок
        title_label = QLabel("👥 Управление аккаунтами пользователей")
        title_label.setStyleSheet("font-size: 14px; font-weight: bold; padding: 10px; background-color: #e8f5e9; color: #000000; border-radius: 5px;")
        layout.addWidget(title_label)

        # Кнопки управления
        buttons_layout = QHBoxLayout()
        
        refresh_btn = QPushButton("🔄 Обновить")
        refresh_btn.clicked.connect(self.load_accounts_data)
        buttons_layout.addWidget(refresh_btn)
        
        send_email_btn = QPushButton("📧 Отправить письмо")
        send_email_btn.clicked.connect(self.send_email_to_employee)
        buttons_layout.addWidget(send_email_btn)
        
        delete_account_btn = QPushButton("🗑️ Удалить аккаунт")
        delete_account_btn.clicked.connect(self.delete_account)
        buttons_layout.addWidget(delete_account_btn)
        
        buttons_layout.addStretch()
        layout.addLayout(buttons_layout)

        # Таблица аккаунтов
        self.accounts_table = QTableView()
        layout.addWidget(self.accounts_table)

    def load_accounts_data(self):
        """Загрузка списка всех аккаунтов"""
        print(" Загрузка аккаунтов...")

        model = QSqlQueryModel()

        query = """
        SELECT 
            u.user_id as 'ID',
            e.last_name || ' ' || e.first_name as 'ФИО',
            u.username as 'Логин',
            u.password as 'Пароль (хеш)',
            u.role as 'Роль',
            CASE u.is_active WHEN 1 THEN 'Активен' ELSE 'Неактивен' END as 'Статус',
            u.created_at as 'Дата создания'
        FROM Users u
        LEFT JOIN Employees e ON u.employee_id = e.employee_id
        ORDER BY u.created_at DESC
        """

        model.setQuery(query, self.db_connection)
        self.accounts_table.setModel(model)
        self.accounts_table.resizeColumnsToContents()

        row_count = model.rowCount()
        print(f" Загружено аккаунтов: {row_count}")

    def delete_account(self):
        """Удаление аккаунта пользователя с подтверждением пароля"""
        # Проверяем, выбрана ли строка
        selected_indexes = self.accounts_table.selectedIndexes()
        if not selected_indexes:
            QMessageBox.warning(self, "Ошибка", "Выберите аккаунт для удаления!")
            return
        
        # Получаем данные выбранной строки
        row = selected_indexes[0].row()
        model = self.accounts_table.model()
        
        user_id = model.data(model.index(row, 0))
        user_name = model.data(model.index(row, 1))
        username = model.data(model.index(row, 2))
        user_role = model.data(model.index(row, 4))
        
        # Нельзя удалить админа
        if user_role == 'admin':
            QMessageBox.warning(self, "Ошибка", "Нельзя удалить аккаунт администратора!")
            return
        
        # Подтверждение удаления
        confirm = QMessageBox.question(
            self,
            "Подтверждение удаления",
            f"Вы уверены, что хотите удалить аккаунт?\n\nФИО: {user_name}\nЛогин: {username}\n\nЭто действие необратимо!",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No
        )
        
        if confirm != QMessageBox.StandardButton.Yes:
            return
        
        # Запрос пароля для подтверждения
        from PyQt6.QtWidgets import QInputDialog, QLineEdit
        password, ok = QInputDialog.getText(
            self,
            "Подтверждение удаления",
            "Введите ваш пароль для подтверждения удаления:",
            QLineEdit.EchoMode.Password
        )
        
        if not ok or not password:
            return
        
        # Проверяем пароль текущего администратора
        import hashlib
        password_hash = hashlib.sha256(password.encode()).hexdigest()
        
        # Специальный случай для admin/admin
        if self.current_user.get('username') == 'admin' and password == 'admin':
            password_valid = True
        else:
            # Проверяем пароль из БД
            check_query = "SELECT password FROM Users WHERE user_id = ?"
            result = self.db.execute_query(check_query, (self.current_user.get('user_id'),))
            password_valid = result and result[0][0] == password_hash
        
        if not password_valid:
            QMessageBox.critical(self, "Ошибка", "Неверный пароль!")
            return
        
        # Удаляем аккаунт
        try:
            delete_query = "DELETE FROM Users WHERE user_id = ?"
            self.db.execute_update(delete_query, (user_id,))
            
            QMessageBox.information(self, "Успех", f"Аккаунт '{username}' успешно удален!")
            print(f"Удален аккаунт: {username} (ID: {user_id})")
            
            # Обновляем таблицу
            self.load_accounts_data()
            
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка при удалении аккаунта:\n{str(e)}")
            print(f"Ошибка удаления аккаунта: {e}")

    def send_email_to_employee(self):
        """Отправка email-уведомления выбранному сотруднику"""
        # Проверяем, выбран ли аккаунт
        current_index = self.accounts_table.currentIndex()
        if not current_index.isValid():
            QMessageBox.warning(self, "Ошибка", "Выберите сотрудника из списка!")
            return
        
        model = self.accounts_table.model()
        row = current_index.row()
        
        # Получаем данные сотрудника
        user_id = model.data(model.index(row, 0))  # ID
        employee_name = model.data(model.index(row, 1))  # ФИО
        username = model.data(model.index(row, 2))  # Логин
        
        # Получаем employee_id и email
        employee_data = self.db.execute_query("""
            SELECT e.employee_id, e.email
            FROM Users u
            JOIN Employees e ON u.employee_id = e.employee_id
            WHERE u.user_id = ?
        """, (user_id,))
        
        if not employee_data:
            QMessageBox.warning(self, "Ошибка", "Не удалось получить данные сотрудника!")
            return
        
        employee_id, email = employee_data[0]
        
        if not email or '@' not in email:
            QMessageBox.warning(
                self, 
                "Ошибка", 
                f"У сотрудника {employee_name} не указан email адрес!\n\n"
                "Добавьте email в профиле сотрудника."
            )
            return
        
        # Проверяем, настроен ли email-notifier
        if not hasattr(self.notification_manager, 'email_notifier') or not self.notification_manager.email_notifier.enabled:
            QMessageBox.warning(
                self,
                "Email не настроен",
                "Email-уведомления не настроены!\n\n"
                "Настройте SMTP-сервер в main.py:\n"
                "self.notification_manager.configure_email(...)"
            )
            return
        
        # Получаем активные выдачи сотрудника
        active_issues = self.db.execute_query("""
            SELECT 
                a.name,
                uh.planned_return_date,
                CAST((julianday(uh.planned_return_date) - julianday('now')) AS INTEGER) as days_until
            FROM Usage_History uh
            JOIN Assets a ON uh.asset_id = a.asset_id
            WHERE uh.employee_id = ?
                AND uh.operation_type = 'выдача'
                AND uh.actual_return_date IS NULL
            ORDER BY uh.planned_return_date ASC
        """, (employee_id,))
        
        if not active_issues:
            QMessageBox.information(
                self,
                "Нет активных выдач",
                f"У сотрудника {employee_name} нет активных выдач инструментов."
            )
            return
        
        # Показываем диалог выбора темы письма
        dialog = QDialog(self)
        dialog.setWindowTitle("Выбор темы письма")
        dialog.setMinimumWidth(500)
        
        dialog_layout = QVBoxLayout(dialog)
        
        # Заголовок
        header = QLabel(f"Отправка письма сотруднику: {employee_name}")
        header.setStyleSheet("font-weight: bold; font-size: 14px; padding: 10px;")
        dialog_layout.addWidget(header)
        
        info_label = QLabel(f"Email: {email}")
        info_label.setStyleSheet("color: #666; padding: 5px;")
        dialog_layout.addWidget(info_label)
        
        # Список активных выдач
        issues_text = "Активные выдачи:\n"
        has_overdue = False
        has_upcoming = False
        
        for asset_name, return_date, days_until in active_issues:
            if days_until < 0:
                status = f"🚨 ПРОСРОЧКА {abs(days_until)} дн."
                has_overdue = True
            elif days_until == 0:
                status = "⚠️ Срок истекает СЕГОДНЯ"
                has_overdue = True
            elif days_until == 1:
                status = "⏰ Срок истекает ЗАВТРА"
                has_upcoming = True
            else:
                status = f"Осталось {days_until} дн."
                has_upcoming = True
            
            issues_text += f"  • {asset_name} — {return_date} ({status})\n"
        
        issues_label = QLabel(issues_text)
        issues_label.setStyleSheet("background-color: #f5f5f5; padding: 10px; border-radius: 5px; font-family: monospace;")
        dialog_layout.addWidget(issues_label)
        
        # Выбор темы письма
        theme_group = QGroupBox("Выберите тему письма:")
        theme_layout = QVBoxLayout()
        
        radio_upcoming = QRadioButton("⏰ Приближается срок сдачи (напоминание)")
        radio_overdue = QRadioButton("🚨 Есть просрочка (срочное уведомление)")
        
        # Выбираем по умолчанию в зависимости от наличия просрочек
        if has_overdue:
            radio_overdue.setChecked(True)
        else:
            radio_upcoming.setChecked(True)
        
        theme_layout.addWidget(radio_upcoming)
        theme_layout.addWidget(radio_overdue)
        theme_group.setLayout(theme_layout)
        dialog_layout.addWidget(theme_group)
        
        # Кнопки
        button_box = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel
        )
        button_box.accepted.connect(dialog.accept)
        button_box.rejected.connect(dialog.reject)
        dialog_layout.addWidget(button_box)
        
        # Показываем диалог
        if dialog.exec() != QDialog.DialogCode.Accepted:
            return
        
        # Определяем тему
        is_overdue = radio_overdue.isChecked()
        
        # Отправляем письма для каждого инструмента
        sent_count = 0
        failed_count = 0
        
        for asset_name, return_date, days_until in active_issues:
            # Фильтруем по выбранной теме
            if is_overdue and days_until >= 1:
                continue  # Пропускаем не просроченные
            if not is_overdue and days_until < 0:
                continue  # Пропускаем просроченные
            
            try:
                success = self.notification_manager.email_notifier.send_deadline_warning(
                    employee_email=email,
                    employee_name=employee_name,
                    asset_name=asset_name,
                    deadline_date=return_date,
                    days_until=days_until
                )
                
                if success:
                    sent_count += 1
                else:
                    failed_count += 1
                    
            except Exception as e:
                print(f"Ошибка отправки письма: {e}")
                failed_count += 1
        
        # Показываем результат
        if sent_count > 0:
            QMessageBox.information(
                self,
                "Письма отправлены",
                f"✅ Отправлено писем: {sent_count}\n"
                f"{'❌ Ошибок: ' + str(failed_count) if failed_count > 0 else ''}\n\n"
                f"Получатель: {employee_name} ({email})"
            )
        else:
            QMessageBox.warning(
                self,
                "Письма не отправлены",
                f"Не найдено инструментов для выбранной темы письма.\n\n"
                f"{'Попробуйте выбрать другую тему.' if failed_count == 0 else 'Произошли ошибки при отправке.'}"
            )

    def setup_user_profile_tab(self):
        """Настройка вкладки профиля пользователя"""
        layout = QVBoxLayout(self.user_profile_tab)
        layout.setSpacing(15)
        layout.setContentsMargins(20, 20, 20, 20)

        # Заголовок
        title_label = QLabel("👤 Мой профиль")
        title_label.setStyleSheet("font-size: 18px; font-weight: bold; padding: 10px; background-color: #e3f2fd; color: #000000; border-radius: 5px;")
        layout.addWidget(title_label)

        # Описание
        info_label = QLabel("Здесь вы можете просмотреть и изменить свою личную информацию. Для сохранения изменений потребуется ввести пароль.")
        info_label.setWordWrap(True)
        info_label.setStyleSheet("color: #666; padding: 5px;")
        layout.addWidget(info_label)

        # Форма с полями
        form_widget = QWidget()
        form_layout = QGridLayout(form_widget)
        form_layout.setSpacing(10)

        # Получаем данные пользователя
        employee_id = self.current_user.get('employee_id')
        if employee_id:
            employee_data = self.db.execute_query("""
                SELECT e.last_name, e.first_name, e.patronymic, 
                       p.position_name, e.phone, e.email
                FROM Employees e
                LEFT JOIN Positions p ON e.position_id = p.position_id
                WHERE e.employee_id = ?
            """, (employee_id,))
            
            if employee_data:
                last_name, first_name, patronymic, position_name, phone, email = employee_data[0]
            else:
                last_name, first_name, patronymic, position_name, phone, email = "", "", "", "", "", ""
        else:
            last_name, first_name, patronymic, position_name, phone, email = "", "", "", "", "", ""

        # Поле: Фамилия
        row = 0
        form_layout.addWidget(QLabel("Фамилия:"), row, 0)
        self.profile_last_name = QLineEdit(last_name if last_name else "")
        self.profile_last_name.setPlaceholderText("Введите фамилию")
        form_layout.addWidget(self.profile_last_name, row, 1)

        # Поле: Имя
        row += 1
        form_layout.addWidget(QLabel("Имя:"), row, 0)
        self.profile_first_name = QLineEdit(first_name if first_name else "")
        self.profile_first_name.setPlaceholderText("Введите имя")
        form_layout.addWidget(self.profile_first_name, row, 1)

        # Поле: Отчество
        row += 1
        form_layout.addWidget(QLabel("Отчество:"), row, 0)
        self.profile_patronymic = QLineEdit(patronymic if patronymic else "")
        self.profile_patronymic.setPlaceholderText("Введите отчество (необязательно)")
        form_layout.addWidget(self.profile_patronymic, row, 1)

        # Поле: Должность (только чтение)
        row += 1
        form_layout.addWidget(QLabel("Должность:"), row, 0)
        self.profile_position = QLineEdit(position_name if position_name else "Не указана")
        self.profile_position.setReadOnly(True)
        self.profile_position.setStyleSheet("background-color: #f0f0f0;")
        form_layout.addWidget(self.profile_position, row, 1)

        # Поле: Телефон
        row += 1
        form_layout.addWidget(QLabel("Телефон:"), row, 0)
        self.profile_phone = QLineEdit(phone if phone else "+7")
        self.profile_phone.setPlaceholderText("+7 (XXX) XXX-XX-XX")
        # Устанавливаем валидатор для телефона
        self.profile_phone.textChanged.connect(self._validate_phone_input)
        form_layout.addWidget(self.profile_phone, row, 1)

        # Поле: Email
        row += 1
        form_layout.addWidget(QLabel("Email:"), row, 0)
        self.profile_email = QLineEdit(email if email else "AlexAndreev132@yandex.ru")
        self.profile_email.setPlaceholderText("example@domain.com")
        form_layout.addWidget(self.profile_email, row, 1)

        layout.addWidget(form_widget)

        # Кнопки
        buttons_layout = QHBoxLayout()
        buttons_layout.addStretch()

        reset_btn = QPushButton("🔄 Сбросить изменения")
        reset_btn.clicked.connect(self.reset_profile_form)
        buttons_layout.addWidget(reset_btn)

        save_btn = QPushButton("💾 Сохранить изменения")
        save_btn.setStyleSheet("background-color: #4CAF50; color: white; font-weight: bold; padding: 10px 20px;")
        save_btn.clicked.connect(self.save_profile_changes)
        buttons_layout.addWidget(save_btn)

        layout.addLayout(buttons_layout)
        layout.addStretch()

    def _validate_phone_input(self, text):
        """Валидация ввода телефона - автоматически добавляет +7"""
        if not text.startswith("+7"):
            self.profile_phone.setText("+7")
            self.profile_phone.setCursorPosition(2)

    def reset_profile_form(self):
        """Сброс формы к исходным значениям"""
        employee_id = self.current_user.get('employee_id')
        if not employee_id:
            return

        employee_data = self.db.execute_query("""
            SELECT e.last_name, e.first_name, e.patronymic, 
                   p.position_name, e.phone, e.email
            FROM Employees e
            LEFT JOIN Positions p ON e.position_id = p.position_id
            WHERE e.employee_id = ?
        """, (employee_id,))

        if employee_data:
            last_name, first_name, patronymic, position_name, phone, email = employee_data[0]
            self.profile_last_name.setText(last_name if last_name else "")
            self.profile_first_name.setText(first_name if first_name else "")
            self.profile_patronymic.setText(patronymic if patronymic else "")
            self.profile_position.setText(position_name if position_name else "Не указана")
            self.profile_phone.setText(phone if phone else "+7")
            self.profile_email.setText(email if email else "AlexAndreev132@yandex.ru")

        QMessageBox.information(self, "Сброшено", "Форма возвращена к исходным значениям")

    def save_profile_changes(self):
        """Сохранение изменений профиля с подтверждением паролем"""
        # Валидация полей
        last_name = self.profile_last_name.text().strip()
        first_name = self.profile_first_name.text().strip()
        patronymic = self.profile_patronymic.text().strip()
        phone = self.profile_phone.text().strip()
        email = self.profile_email.text().strip()

        # Проверка обязательных полей
        if not last_name:
            QMessageBox.warning(self, "Ошибка", "Фамилия не может быть пустой")
            return

        if not first_name:
            QMessageBox.warning(self, "Ошибка", "Имя не может быть пустым")
            return

        # Валидация: только буквы в ФИО
        if not last_name.replace('-', '').isalpha():
            QMessageBox.warning(self, "Ошибка", "Фамилия должна содержать только буквы")
            return

        if not first_name.replace('-', '').isalpha():
            QMessageBox.warning(self, "Ошибка", "Имя должно содержать только буквы")
            return

        if patronymic and not patronymic.replace('-', '').isalpha():
            QMessageBox.warning(self, "Ошибка", "Отчество должно содержать только буквы")
            return

        # Валидация телефона
        if not phone.startswith("+7"):
            QMessageBox.warning(self, "Ошибка", "Номер телефона должен начинаться с +7")
            return

        if len(phone) < 12:
            QMessageBox.warning(self, "Ошибка", "Введите корректный номер телефона")
            return

        # Валидация email
        if email and '@' not in email:
            QMessageBox.warning(self, "Ошибка", "Введите корректный email адрес")
            return

        # Подтверждение изменений
        confirm = QMessageBox.question(
            self,
            "Подтверждение",
            f"Сохранить следующие изменения?\n\n"
            f"ФИО: {last_name} {first_name} {patronymic}\n"
            f"Телефон: {phone}\n"
            f"Email: {email}",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )

        if confirm != QMessageBox.StandardButton.Yes:
            return

        # Запрос пароля для подтверждения
        password, ok = QInputDialog.getText(
            self,
            "Подтверждение пароля",
            "Введите ваш пароль для подтверждения изменений:",
            QLineEdit.EchoMode.Password
        )

        if not ok or not password:
            return

        # Проверка пароля
        import hashlib
        password_hash = hashlib.sha256(password.encode()).hexdigest()
        
        user_id = self.current_user.get('user_id')
        stored_hash = self.db.execute_query(
            "SELECT password FROM Users WHERE user_id = ?",
            (user_id,)
        )

        if not stored_hash or stored_hash[0][0] != password_hash:
            QMessageBox.critical(self, "Ошибка", "Неверный пароль!")
            return

        # Сохранение изменений
        try:
            employee_id = self.current_user.get('employee_id')
            
            # Применяем capitalize()
            last_name = last_name.capitalize()
            first_name = first_name.capitalize()
            patronymic = patronymic.capitalize() if patronymic else None

            self.db.execute_update("""
                UPDATE Employees 
                SET last_name = ?, first_name = ?, patronymic = ?, phone = ?, email = ?
                WHERE employee_id = ?
            """, (last_name, first_name, patronymic, phone, email, employee_id))

            # Обновляем текущего пользователя
            full_name = f"{last_name} {first_name}"
            if patronymic:
                full_name += f" {patronymic}"
            
            self.current_user['full_name'] = full_name
            
            # Обновляем метку информации о пользователе
            self.user_info_label.setText(f"👤 Вы вошли как: {full_name} ({self.current_user.get('role', 'user').upper()})")

            QMessageBox.information(self, "Успех", "✅ Изменения успешно сохранены!")
            
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка при сохранении:\n{str(e)}")
            print(f"Ошибка сохранения профиля: {e}")

    def load_requests_data(self):
        """Загрузка списка запросов на выдачу активов"""
        print(" Загрузка запросов на выдачу активов...")

        if not hasattr(self, 'db_connection'):
            self.db_connection = QSqlDatabase.addDatabase("QSQLITE")
            self.db_connection.setDatabaseName("inventory.db")

        if not self.db_connection.isOpen():
            if not self.db_connection.open():
                return

        # Очищаем старую модель
        if hasattr(self, 'requests_table') and self.requests_table.model():
            self.requests_table.setModel(None)

        model = QSqlQueryModel()

        query = """
        SELECT 
            ar.request_id as 'ID',
            a.name as 'Актив',
            a.model as 'Модель',
            e.last_name || ' ' || e.first_name as 'Сотрудник',
            ar.request_date as 'Дата запроса',
            ar.planned_return_date as 'План возврата',
            CASE ar.status
                WHEN 'pending' THEN 'Ожидает'
                WHEN 'approved' THEN 'Одобрено'
                WHEN 'rejected' THEN 'Отклонено'
                ELSE ar.status
            END as 'Статус',
            COALESCE(ar.notes, '') as 'Примечание'
        FROM Asset_Requests ar
        JOIN Assets a ON ar.asset_id = a.asset_id
        JOIN Employees e ON ar.employee_id = e.employee_id
        ORDER BY ar.request_date DESC
        """

        model.setQuery(query, self.db_connection)
        
        # Устанавливаем модель только если таблица существует (только для админов)
        if hasattr(self, 'requests_table'):
            self.requests_table.setModel(model)
            self.requests_table.resizeColumnsToContents()

        row_count = model.rowCount()
        print(f" Загружено запросов: {row_count}")

    def approve_request(self):
        """Одобрение запроса на выдачу актива"""
        current_index = self.requests_table.currentIndex()
        if not current_index.isValid():
            QMessageBox.warning(self, "Ошибка", "Выберите запрос из таблицы!")
            return

        model = self.requests_table.model()
        row = current_index.row()

        request_id = model.data(model.index(row, 0))
        asset_name = model.data(model.index(row, 1))
        employee_name = model.data(model.index(row, 3))

        # Подтверждение
        reply = QMessageBox.question(
            self,
            "Подтверждение",
            f"Одобрить запрос на выдачу '{asset_name}' для {employee_name}?"
        )

        if reply != QMessageBox.StandardButton.Yes:
            return

        try:
            # Получаем информацию из запроса
            req_data = self.db.execute_query(
                "SELECT asset_id, employee_id, planned_return_date, notes FROM Asset_Requests WHERE request_id = ?",
                (request_id,)
            )[0]

            asset_id, employee_id, planned_return_date, notes = req_data

            # Проверяем количество доступных активов
            asset_data = self.db.execute_query(
                "SELECT quantity, current_status FROM Assets WHERE asset_id = ?",
                (asset_id,)
            )[0]

            current_qty, status = asset_data

            if current_qty <= 0:
                QMessageBox.warning(self, "Ошибка", "На складе нет доступных единиц!")
                return

            # Создаем операцию выдачи
            history_query = """
            INSERT INTO Usage_History (asset_id, employee_id, operation_type, operation_date, planned_return_date, notes)
            VALUES (?, ?, ?, ?, ?, ?)
            """

            self.db.execute_update(
                history_query,
                (asset_id, employee_id, 'выдача', datetime.now().isoformat(), planned_return_date, notes)
            )

            # Обновляем количество активов
            new_qty = current_qty - 1

            self.db.execute_update(
                "UPDATE Assets SET quantity = ? WHERE asset_id = ?",
                (new_qty, asset_id)
            )
            
            # Обновляем статус на основе активных выдач
            self.db.update_asset_status(asset_id)

            # Обновляем статус запроса
            self.db.execute_update(
                "UPDATE Asset_Requests SET status = ?, approved_by = ?, approved_at = ? WHERE request_id = ?",
                ('approved', int(self.current_user.get('user_id', 0)), datetime.now().isoformat(), request_id)
            )

            QMessageBox.information(self, "Успех", "✅ Запрос одобрен и актив выдан!")
            self._refresh_all_data()

        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка при одобрении запроса:\n{str(e)}")

    def reject_request(self):
        """Отклонение запроса на выдачу актива"""
        current_index = self.requests_table.currentIndex()
        if not current_index.isValid():
            QMessageBox.warning(self, "Ошибка", "Выберите запрос из таблицы!")
            return

        model = self.requests_table.model()
        row = current_index.row()

        request_id = model.data(model.index(row, 0))
        asset_name = model.data(model.index(row, 1))

        # Подтверждение
        reply = QMessageBox.question(
            self,
            "Подтверждение",
            f"Отклонить запрос на выдачу '{asset_name}'?"
        )

        if reply != QMessageBox.StandardButton.Yes:
            return

        try:
            # Обновляем статус запроса
            self.db.execute_update(
                "UPDATE Asset_Requests SET status = ?, approved_by = ?, approved_at = ? WHERE request_id = ?",
                ('rejected', int(self.current_user.get('user_id', 0)), datetime.now().isoformat(), request_id)
            )

            QMessageBox.information(self, "Успех", "✅ Запрос отклонен!")
            self.load_requests_data()

        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка при отклонении запроса:\n{str(e)}")

    def request_asset(self):
        """Создание запроса на выдачу актива (для обычных пользователей)"""
        if self.current_user.get('role') not in ['user', 'admin']:
            QMessageBox.warning(self, "Ошибка", "У вас нет прав для создания запроса!")
            return

        dialog = RequestAssetDialog(self.current_user, self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            # Обновляем список запросов только если таблица существует (для админов)
            if hasattr(self, 'requests_table'):
                self.load_requests_data()

    def logout(self):
        """Выход и возврат на экран авторизации"""
        reply = QMessageBox.question(
            self,
            "Выход",
            "Вы уверены, что хотите выйти?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            self.close()

    def closeEvent(self, event):
        """Обработчик закрытия окна приложения"""
        print("Закрытие окна...")
        
        # Останавливаем проверку уведомлений
        if hasattr(self, 'notification_manager'):
            self.notification_manager.cleanup()
        
        super().closeEvent(event)


def main():
    print("Запуск приложения...")
    app = QApplication(sys.argv)
    
    # Устанавливаем иконку приложения (для Dock/панели задач)
    app_icon_path = Path("pictures/InstrumentTracker_icon.png")
    if app_icon_path.exists():
        app.setWindowIcon(QIcon(str(app_icon_path)))
        print(f" Установлена иконка приложения: {app_icon_path}")
    else:
        print(f"️ Иконка приложения не найдена: {app_icon_path}")
    
    while True:
        # Показываем окно входа
        login_dialog = LoginDialog()
        current_user = None
        
        def on_login_success(user_data):
            nonlocal current_user
            current_user = user_data
            # Логируем вход
            AuditLogger.log_action(
                user_data.get('user_id'),
                user_data.get('username'),
                'user_login',
                {'role': user_data.get('role')}
            )
            login_dialog.accept()
        
        login_dialog.login_successful.connect(on_login_success)
        
        if login_dialog.exec() == QDialog.DialogCode.Accepted and current_user:
            # Создаем главное окно с текущим пользователем
            window = MainWindow(current_user)
            
            # Устанавливаем иконку окна (логотип)
            logo_path = Path("pictures/InstrumentTracker_logo.png")
            if logo_path.exists():
                window.setWindowIcon(QIcon(str(logo_path)))
                print(f" Установлен логотип окна: {logo_path}")
            else:
                print(f"️ Логотип окна не найден: {logo_path}")
            
            window.show()
            print(" Приложение запущено успешно")
            app.exec()  # Ждем закрытия окна
            # После закрытия окна, цикл продолжится и покажет логин снова
        else:
            print(" Вход отменен")
            break
    
    sys.exit(0)


if __name__ == "__main__":
    main()
